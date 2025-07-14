import streamlit as st
import pandas as pd
from datetime import date
import calendar 
from calendar import monthrange
import jpholiday
import numpy as np
from ortools.sat.python import cp_model
import optimizer
import copy
import io
from openpyxl.styles import PatternFill
import pandas as pd
import matplotlib.pyplot as plt
import datetime

st.set_page_config(layout="wide")
st.title("🗓️ スマートシフト作成アプリ")

# タブ分割
tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs([
    "🗓️ 年月設定", "⏰ シフト設定", "👥 スタッフ設定", "🔢 必要人数入力",
    "📝 希望休・有給",  "🚀 シフト作成・最適化"
])

# グローバル変数的なもの（session_stateに保存）
if "staff_names" not in st.session_state:
    st.session_state.staff_names = []
if "shifts" not in st.session_state:
    st.session_state.shifts = [{"label": "日勤", "hours": 8}]

def to_colored_excel(df, year=None):

    used_shifts = set()
    for col in df.columns[:-1]:  # 最後の列は「総勤務時間」など集計なので除外
        used_shifts.update(df[col].dropna().unique())

    color_map = {
        "休み": "FFFF00",
        "応援": "FFC0CB",
        "有給": "90EE90",
    }

    shifts_to_color = [sh for sh in used_shifts if sh not in color_map]
    cmap = plt.get_cmap("tab20")
    for i, shift in enumerate(shifts_to_color):
        rgb = tuple(int(255 * x) for x in cmap(i % 20)[:3])
        hex_color = '{:02X}{:02X}{:02X}'.format(*rgb)
        color_map[shift] = hex_color

    if year is None:
        year = datetime.datetime.now().year

    # --- シフト回数集計表の作成 ---
    shift_labels = list(color_map.keys())
    shift_counts_df = pd.DataFrame(index=df.index, columns=shift_labels).fillna(0)

    for staff in df.index:
        for day in df.columns:
            if day == "総勤務時間":
                continue
            shift = df.at[staff, day]
            if shift in shift_labels:
                shift_counts_df.at[staff, shift] += 1

    shift_counts_df["総勤務時間"] = df["総勤務時間"]

    # --- Excel出力 ---
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=True, sheet_name="シフト結果")
        ws = writer.sheets["シフト結果"]

        # 色付け
        for row in range(2, len(df) + 2):
            for col in range(2, len(df.columns) + 2):
                cell = ws.cell(row=row, column=col)
                shift = cell.value
                if shift in color_map:
                    fill = PatternFill(start_color=color_map[shift], end_color=color_map[shift], fill_type="solid")
                    cell.fill = fill

        for col_idx, col_name in enumerate(df.columns[:-1], start=2):
            try:
                month, day = map(int, col_name.split("/"))
                date = datetime.date(year, month, day)
                if date.weekday() >= 5 or jpholiday.is_holiday(date):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
            except Exception:
                pass

        # --- 📊 シフト割当数テーブルを df の下に追加 ---
        start_row = len(df) + 4  # 1行空けてから書き始め
        shift_counts_df.to_excel(writer, sheet_name="シフト結果", startrow=start_row, index=True)

    return output.getvalue()

# 🗓️ 年月設定
with tab1:
    st.subheader("対象年月を選択してください")
    year = st.selectbox("年", range(2024, 2031), index=1)
    month = st.selectbox("月", range(1, 13), index=date.today().month - 1)
    num_days = monthrange(year, month)[1]
    dates = [date(year, month, d+1) for d in range(num_days)]
    st.session_state["dates"] = dates
    st.success(f"{year}年{month}月（{num_days}日間）が対象です。")

with tab2:
    st.header("👥 スタッフ設定")

    # 月の平日数計算
    dates = st.session_state["dates"]
    weekdays = [d.weekday() for d in dates]
    num_weekdays = sum(
        1 for d in dates if d.weekday() < 5 and not jpholiday.is_holiday(d))
    st.info(f"この月の平日数は **{num_weekdays}日** です。")

    num_staff = st.number_input("スタッフ数", 1, 20, 5)

    staff_names = []
    daily_work_hours_dict = {}
    total_work_hours = {}

    for i in range(num_staff):
        with st.container():
            cols = st.columns([2, 2, 3, 3])  # 横4分割：名前・1日の勤務時間・月総勤務時間・対応シフト
            with cols[0]:
                name = st.text_input(f"スタッフ{i+1}の名前", value=f"スタッフ{i+1}", key=f"name_{i}")
                staff_names.append(name)
            with cols[1]:
                daily_hours = st.number_input(f"１日の勤務時間（時間）", min_value=0.0, value=8.0, step=0.5, key=f"daily_hours_{i}")
                daily_work_hours_dict[name] = daily_hours
            with cols[2]:
                monthly_hours = daily_hours * num_weekdays
                st.markdown(f"月の総勤務時間（平日ベース）: **{monthly_hours:.1f}時間**")
                total_work_hours[name] = monthly_hours
            with cols[3]:
                st.markdown("**対応可能なシフト**")
                compatible_shifts = []
                for shift in [s["label"] for s in st.session_state.shifts]:
                    checked = st.checkbox(
                        shift,
                        value=True,
                        key=f"compatible_{i}_{shift}"
                    )
                    if checked:
                        compatible_shifts.append(shift)
                # スタッフ名をキーにして保存（nameはこのループ内で定義済み）
                st.session_state.setdefault("shift_compatibility", {})
                st.session_state["shift_compatibility"][name] = compatible_shifts
    
            st.markdown("---")  # 区切り線

    st.session_state["staff_names"] = staff_names
    st.session_state["daily_work_hours"] = daily_work_hours_dict
    st.session_state["total_work_hours"] = total_work_hours

with tab3:
    st.subheader("⏰シフト種別の設定　（保存忘れずに）")

    FIXED_SHIFT = {"label": "日勤", "hours": 8}  # 固定シフト

    # 初期化時、日勤を必ず登録
    if "shifts" not in st.session_state:
        st.session_state.shifts = [FIXED_SHIFT]

    # 編集用バッファは日勤以外
    if "shift_edit_buffer" not in st.session_state:
        st.session_state.shift_edit_buffer = [s for s in st.session_state.shifts if s["label"] != FIXED_SHIFT["label"]]

    buffer = st.session_state.shift_edit_buffer

    # 固定シフトを表示（編集不可）
    st.markdown("### 固定シフト")
    st.write(f"**{FIXED_SHIFT['label']}**  —  {FIXED_SHIFT['hours']}時間 （編集・削除不可）")

    st.markdown("---")

    with st.form("shift_form"):
        st.write("※ 編集後、「保存」または「削除」ボタンで操作を確定してください")
        new_buffer = []
        delete_flags = []

        for i, s in enumerate(buffer):
            cols = st.columns([4, 2, 1])
            label = cols[0].text_input(f"label_{i}", value=s["label"], key=f"edit_label_{i}")
            hours = cols[1].number_input(f"hours_{i}", value=s["hours"], min_value=1, max_value=24, key=f"edit_hours_{i}")
            delete = cols[2].checkbox("削除", key=f"delete_{i}")
            delete_flags.append(delete)
            new_buffer.append({"label": label, "hours": hours})

        col1, col2, col3 = st.columns(3)
        add_btn = col1.form_submit_button("➕ シフト追加")
        save_btn = col2.form_submit_button("💾 保存")
        delete_btn = col3.form_submit_button("🗑️ チェックしたシフトを削除")

        if add_btn:
            new_buffer.append({"label": f"シフト{len(new_buffer)+1}", "hours": 8})
            st.session_state.shift_edit_buffer = new_buffer
            st.rerun()

        if save_btn:
            # 日勤を必ず先頭にして保存
            combined = [FIXED_SHIFT] + new_buffer
            st.session_state.shift_edit_buffer = new_buffer
            st.session_state.shifts = combined
            st.success("✅ シフト情報が保存されました")
            st.rerun()

        if delete_btn:
            # 削除チェックされてないものだけ残す
            new_buffer = [s for s, d in zip(new_buffer, delete_flags) if not d]
            st.session_state.shift_edit_buffer = new_buffer
            st.session_state.shifts = [FIXED_SHIFT] + new_buffer
            st.success("🗑️ 選択したシフトを削除しました")
            st.rerun()


# 🔢 必要人数入力
with tab4:
    st.subheader("📅 必要人数（保存忘れずに）")

    year = year
    month = month
    month_weeks = calendar.monthcalendar(year, month)
    shift_types = [s["label"] for s in st.session_state.shifts]

    if "temp_required_staff" not in st.session_state:
        st.session_state["temp_required_staff"] = {}

    st.markdown("### ⚙️ 一括設定（全日・全シフト）")

    batch_input = {}
    cols = st.columns(len(shift_types))
    for i, shift in enumerate(shift_types):
        batch_input[shift] = cols[i].number_input(
            f"{shift}",
            min_value=0,
            max_value=20,
            value=0,
            key=f"batch_{shift}"
        )

    if st.button("一括反映"):
        for week in month_weeks:
            for day in week:
                if day == 0:
                    continue
                d = date(year, month, day)
    
                if d not in st.session_state["temp_required_staff"]:
                    st.session_state["temp_required_staff"][d] = {}
    
                for shift in shift_types:
                    # 土日祝も含めて反映（0にしない）
                    st.session_state["temp_required_staff"][d][shift] = batch_input[shift]
    
        st.success("一括設定を全日付に反映しました。")

    st.markdown("### 📅 日毎の必要人数を調整（編集後、保存ボタンで確定）")
    
    # フォームで囲むことで「保存」まで値の反映を遅延
    with st.form("required_staff_form"):
        for week in month_weeks:
            day_cols = st.columns(7)
            for i, day in enumerate(week):
                if day == 0:
                    day_cols[i].empty()
                    continue
    
                d = date(year, month, day)
                weekday = d.weekday()
                is_holiday = jpholiday.is_holiday(d)
    
                # 文字色指定
                if is_holiday:
                    font_color = "deeppink"  # 祝日はピンク文字
                elif weekday == 5:
                    font_color = "blue"      # 土曜は青文字
                elif weekday == 6:
                    font_color = "red"       # 日曜は赤文字
                else:
                    font_color = "black"     # 平日は黒
    
                date_str = f"{d.strftime('%-m/%-d')}（{calendar.day_abbr[weekday]}）"
    
                with day_cols[i]:
                    st.markdown(
                        f'<span style="color:{font_color}; font-weight:bold;">{date_str}</span>',
                        unsafe_allow_html=True
                    )

                    if d not in st.session_state["temp_required_staff"]:
                        st.session_state["temp_required_staff"][d] = {shift: 0 for shift in shift_types}

                    for shift in shift_types:
                        key = f"temp_req_{d}_{shift}"
                        current_val = st.session_state["temp_required_staff"][d].get(shift, 0)

                        new_val = st.number_input(
                            f"{shift}",
                            min_value=0,
                            max_value=20,
                            value=current_val,
                            key=key
                        )
                        # ここでは値を直接書き換えず、保存ボタン押すまで待つ
                     # ✅ 人数固定チェックボックス追加
                    strict_key = f"strict_{d}"
                    strict_val = st.session_state.get("strict_staffing_days", {}).get(d, False)
                    strict_checked = st.checkbox("人数固定", value=strict_val, key=strict_key)

        submitted = st.form_submit_button("💾 必要人数を保存")
        if submitted:
            # フォームの値は st.session_state にすでに入ってるのでコピーするだけ
            temp = {}
            for week in month_weeks:
                for day in week:
                    if day == 0:
                        continue
                    d = date(year, month, day)
                    temp[d] = {}
                    for shift in shift_types:
                        key = f"temp_req_{d}_{shift}"
                        temp[d][shift] = st.session_state.get(key, 0)
            st.session_state["temp_required_staff"] = temp
            st.session_state["required_staff"] = copy.deepcopy(temp)
            # ✅ 人数固定データ保存
            strict_days = {}
            for week in month_weeks:
                for day in week:
                    if day == 0:
                        continue
                    d = date(year, month, day)
                    strict_days[d] = st.session_state.get(f"strict_{d}", False)
            st.session_state["strict_staffing_days"] = strict_days
        
            st.success("必要人数を保存しました。")

# 📝 希望休・有給
with tab5:
    st.subheader("📝 希望入力（保存忘れずに）")

    shift_labels = ["－", "希望休", "有給"] + [s["label"] for s in st.session_state.shifts]
    month_weeks = calendar.monthcalendar(year, month)

    if "leave_requests" not in st.session_state:
        st.session_state["leave_requests"] = {}

    # フォーム内にまとめて保存制御
    with st.form("leave_requests_form"):
        temp_leave_requests = {}

        for name in st.session_state.staff_names:
            st.markdown(f"### 👤 {name}")
            selection = {}

            for week in month_weeks:
                cols = st.columns(7)
                for i, day in enumerate(week):
                    if day == 0:
                        cols[i].empty()
                        continue

                    d = date(year, month, day)
                    label_color = (
                        "red" if d.weekday() == 6 or jpholiday.is_holiday(d) else
                        "blue" if d.weekday() == 5 else "black"
                    )
                    label = f"<span style='color:{label_color}'>{d.day}</span>"

                    with cols[i]:
                        st.markdown(label, unsafe_allow_html=True)
                        key = f"{name}_{d}"
                        sel = st.selectbox(" ", shift_labels, key=key)
                        selection[d] = sel

            temp_leave_requests[name] = {
                "希望休": [d for d, v in selection.items() if v == "希望休"],
                "有給": [d for d, v in selection.items() if v == "有給"],
                "シフト希望": {d: v for d, v in selection.items() if v not in ["－", "希望休", "有給"]}
            }

        submitted = st.form_submit_button("💾 希望休・有給を保存")
        if submitted:
            st.session_state["leave_requests"] = temp_leave_requests
            st.success("希望休・有給・シフト希望を保存しました！")

with tab6:
    st.header("🧠 シフト作成・最適化")

    
    st.session_state["use_support_shift"] = st.checkbox(
        "🛠️ 応援を活用する", value=st.session_state.get("use_support_shift", False)
    )


    # 1. 各種制約の重みをスライダーで調整
    st.subheader("⚙️ ソフト制約の重み設定")
    support_penalty = st.slider("🆘 応援使用許容度（高いほど最低限の応援", 0, 5000, 1000, step=100)
    shift_compat_penalty = st.slider("❌ 対応不可シフト許容度（高いほど可能なシフト以外割り当てられない", 0, 1000, 50, step=10)
    workload_diff_penalty = st.slider("📉 月の労働時間に近づける（基本はMAX", 0, 5000, 50, step=100)
    day_shift_bonus = st.slider("☀️ 余ったシフトを日勤へ誘導(低いほど日勤者増", -500, 0, -50, step=10)

    st.session_state["penalties"] = {
        "support_penalty": support_penalty,
        "shift_compat_penalty": shift_compat_penalty,
        "workload_diff_penalty": workload_diff_penalty,
        "day_shift_bonus": day_shift_bonus,
    }

    
    col1, col2, col3 = st.columns([1, 1, 1])
    trigger = False

    if col1.button("⬅️ 前のシフト案を見る"):
        if st.session_state.solution_index > 0:
            st.session_state.solution_index -= 1
            trigger = True
        else:
            st.warning("これが最初のシフト案です。")

    if col2.button("🚀 最初のシフトを作成"):
        st.session_state.solution_index = 0
        trigger = True

    if col3.button("➡️ 次のシフト案を見る"):
        st.session_state.solution_index += 1
        trigger = True

    if trigger:
        with st.spinner("シフトを最適化中..."):
            result = optimizer.optimize_shifts(
                staff_names=st.session_state["staff_names"],
                shifts=st.session_state["shifts"],
                dates=st.session_state["dates"],
                required_staff=st.session_state["required_staff"],
                leave_requests=st.session_state["leave_requests"],
                daily_work_hours=st.session_state["daily_work_hours"],
                use_support_shift=st.session_state["use_support_shift"],
                total_work_hours=st.session_state["total_work_hours"],
                shift_compatibility=st.session_state.get("shift_compatibility"),
                strict_staffing_days=st.session_state.get("strict_staffing_days"),
                solution_index=st.session_state.solution_index,
                max_solutions=10,
                penalties=st.session_state["penalties"]
            )

        if result is None:
            st.warning("これ以上のシフト案は見つかりません。最初の解に戻してください。")
            st.session_state.solution_index = 0
        else:
            st.session_state["latest_result"] = result

            df_result = pd.DataFrame(result).T
            df_result.columns = [
                d.strftime("%m/%d") if hasattr(d, "strftime") else str(d)
                for d in df_result.columns
            ]

            shift_hours_map = {s["label"]: s["hours"] for s in st.session_state["shifts"]}
            total_hours = []
            for staff in df_result.index:
                total = 0
                daily_hour = st.session_state["daily_work_hours"].get(staff, 0)
                for day in df_result.columns:
                    shift = df_result.at[staff, day]
                    if shift == "有給":
                        total += daily_hour
                    else:
                        total += shift_hours_map.get(shift, 0)
                total_hours.append(total)

            df_result["総勤務時間"] = total_hours
            st.session_state["latest_df_result"] = df_result

        if "latest_df_result" in st.session_state:
            df_result = st.session_state["latest_df_result"].copy()
            shift_labels = [s["label"] for s in st.session_state["shifts"]] + ["休み", "有給", "応援"]
        
            # シフト別集計（総勤務時間列を除外）
            shift_counts_df = pd.DataFrame(index=df_result.index, columns=shift_labels).fillna(0)
        
            for staff in df_result.index:
                for day in df_result.columns:
                    if day == "総勤務時間":
                        continue
                    shift = df_result.at[staff, day]
                    if shift in shift_counts_df.columns:
                        shift_counts_df.at[staff, shift] += 1
        
            shift_counts_df["総勤務時間"] = df_result["総勤務時間"]
            
            shift_counts_df = shift_counts_df.astype(int)
            st.session_state["shift_counts_df"] = shift_counts_df  # ← 追加
        
    # ✅ 表示（セッションに保存された結果を使う）
    if "latest_df_result" in st.session_state:
        st.success(f"✅ 解 #{st.session_state.solution_index + 1} を表示中")
        st.dataframe(st.session_state["latest_df_result"], use_container_width=True)

        if "shift_counts_df" in st.session_state:
            st.markdown("### 📊 シフト割当数（スタッフ別）")
            st.dataframe(st.session_state["shift_counts_df"], use_container_width=True)
        
        # Excelダウンロード用
        excel_data = to_colored_excel(st.session_state["latest_df_result"])

        # 年月のフォールバック（未定義時用）
        year = st.session_state.get("selected_year", 2025)
        month = st.session_state.get("selected_month", 7)

        st.download_button(
            label="📥 色付きシフト表をExcelでダウンロード",
            data=excel_data,
            file_name=f"shift_result_{year}_{month}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
